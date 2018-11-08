'''
Created on 06.06.2018

@author: feuk8fs
'''
import boto3
from botocore.config import Config

AD_USERS_FILE_NAME = 'awsgroupexport.xlsx'


def find_between(s, first, last):
    """
    >>> find_between("abcd", "b", "d")
    c
    """
    try:
        start = s.index(first) + len(first)
        end = s.index(last, start)
        return s[start:end]
    except ValueError:
        return ""


def get_username_password():
    ''' input username and password from the user interface '''
    import utils
    import getpass
    username, password = utils.input_username_password(getpass.getuser())
    username = "win" + '\\' + username
    return username, password


def get_boto_session(username, password, account_id, role):
    ''' login AWS Account (account_id) using username, password and the role, than get the token, than create boto3 session '''
    from pyawslogin import get_assertion, get_token, get_roles
    import sys
    import os
    import tempfile
    # login AWS using username and password and get the assertion
    assertion = get_assertion(username, password, "https://cloudsignin.win.azd.cloud.allianz/adfs/ls/IdpInitiatedSignOn.aspx?loginToRp=urn:amazon:webservices", "/etc/ssl/certs/saml.pem", True, os.path.join(tempfile.gettempdir(), 'saml_cookie.txt'))
    # get all roles that the user has
    roles = get_roles(assertion)
    # get selcted role arn  and principal arn
    for awsrole in roles:
        if account_id in awsrole and role in awsrole:
            selected_role = awsrole
            break
    # get token object
    token = get_token(selected_role, assertion, 'eu-central-1')
    if not token:
        sys.exit('you are not authorized to login to this account:' + account_id)
    # create boto session using the credentials
    session = boto3.Session(
            aws_access_key_id = token['Credentials']['AccessKeyId'],
            aws_secret_access_key = token['Credentials']['SecretAccessKey'],
            aws_session_token = token['Credentials']['SessionToken'],
        )
    return session


def get_iam_client(username, password, account_id, role):
    ''' create iam client '''
    session = get_boto_session(username, password, account_id, role)
    iam_client = session.client('iam')
    return iam_client

#####################################################


def get_account_info(iam_client, account_id):
    ''' using iam client get the name of the current  logged-in account
        OUTPUT account_info = {"account_id": account_id, "account_name": account_name}
    '''
    account_name = ""
    # get all aliases of the current account
    response = iam_client.list_account_aliases(MaxItems = 1)
    aliases = response['AccountAliases']
    # get just the first alias
    if len(aliases) > 0:
        account_name = aliases[0]
    # create account info dictionary
    account_info = {"account_id": account_id, "account_name": account_name}
    return account_info


def get_roles(iam_client):
    '''using boto3 get all roles of the current logged-in account
        OUTPUT: roles =  [ { 'Path': 'string', 'RoleName': 'string', 'RoleId': 'string',  'Arn': 'string',  'CreateDate': datetime(2015, 1, 1),
                              'AssumeRolePolicyDocument': 'string', 'Description': 'string', 'MaxSessionDuration': 123 },],
    '''
    response = iam_client.list_roles()
    roles = response['Roles']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    # if is_truncated is true, that's mean there are more values
    while is_truncated:
        response = iam_client.list_roles(Marker = marker)
        roles.extend(response['Roles'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return roles


def get_wusers(iam_client):
    '''using boto3 get all IAM users of the current logged-in account
        OUTPUT: users =  [ {'Path': 'string', 'UserName': 'string', 'UserId': 'string', 'Arn': 'string', 'CreateDate': datetime(2015, 1, 1), 'PasswordLastUsed': datetime(2015, 1, 1) },]
    '''
    response = iam_client.list_users()
    users = response['Users']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_users(Marker = marker)
        users.extend(response['Users'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return users


def get_groups(iam_client):
    '''using boto3 get all groups of the current logged-in account
        output: groups = [{'Path': 'string', 'GroupName': 'string', 'GroupId': 'string', 'Arn': 'string', 'CreateDate': datetime(2015, 1, 1)},]
    '''
    response = iam_client.list_groups()
    groups = response['Groups']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_groups(Marker = marker)
        groups.extend(response['Groups'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return groups

#############################################################################


def get_mpolicies(iam_client):
    '''using boto3 get all managed policies of the current logged-in account
        output:  policies = [{'PolicyName': 'string','PolicyId': 'string','Arn': 'string',
        'Path': 'string','DefaultVersionId': 'string','AttachmentCount': 123, 'IsAttachable': True|False, 'Description': 'string','CreateDate': datetime(2015, 1, 1), 'UpdateDate': datetime(2015, 1, 1) },]
    '''
    response = iam_client.list_policies()
    policies = response['Policies']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_policies(Marker = marker)
        policies.extend(response['Policies'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return policies


def get_action_from_policy_document(policy_document, policy_id):
    '''using boto3 get all actions from the policy_document and put each action and the policy_id in one dictionary, than return list of actions
        policy_document = {'Statement': [{'action': [....], 'Resource':[...], 'Condition':..., 'Effect': ..., 'Sid': ....}]}
    '''
    # policy_document = json.dumps(policy_document)
    # get statements
    statements = policy_document['Statement']
    actions_list = []
    # because statements is sometimes list and sometime just a dictionary
    if  type(statements) is list:
        # loop over all statements and extract the actions
        for statement in statements:
            resource = str(statement['Resource'])
            effect = str(statement['Effect'])
            condition = str(statement['Condition']) if 'Condition' in statement.keys() else None
            sid = str(statement['Sid']) if 'Sid' in statement.keys() else None
            for key, value in statement.items():
                if key != 'Resource' and key != 'Effect' and key != 'Sid' and key != 'Condition':
                    if type(value) is list:
                        for action in value:
                            action_instance = {'type': key, 'resource': resource, 'effect':effect, 'action': action, 'sid' : sid, 'condition': condition}
                            action_instance.update(policy_id)
                            actions_list.append(action_instance)
                    else:
                        action_instance = {'type': key, 'resource': resource, 'effect':effect, 'action': value, 'sid' : sid, 'condition': condition}
                        action_instance.update(policy_id)
                        actions_list.append(action_instance)
    else:
        # extract the actions
        resource = str(statements['Resource'])
        effect = str(statements['Effect'])
        condition = str(statements['Condition']) if 'Condition' in statements.keys() else None
        sid = str(statements['Sid']) if 'Sid' in statements.keys() else None
        for key, value in statements.items():
            if key != 'Resource' and key != 'Effect' and key != 'Sid' and key != 'Condition':
                if type(value) is list:
                    for action in value:
                        action_instance = {'type': key, 'resource': resource, 'effect':effect, 'action': action, 'sid' : sid, 'condition': condition}
                        action_instance.update(policy_id)
                        actions_list.append(action_instance)
                else:
                    action_instance = {'type': key, 'resource': resource, 'effect':effect, 'action': value, 'sid' : sid, 'condition': condition}
                    action_instance.update(policy_id)
                    actions_list.append(action_instance)
    return actions_list


def get_mpolicy_actions(iam_client, mpolicy):
    ''' get actions of one managed policy using boto get_policy_version
        OUTPUT:  actions_list = [{'policy_id': policy_id, 'type': key, 'resource': resource, 'effect':effect, 'action': value, 'sid' : sid, 'condition': condition}]
    '''
    response = iam_client.get_policy_version(PolicyArn = mpolicy['Arn'], VersionId = mpolicy['DefaultVersionId'])
    policy_version = response['PolicyVersion']
    policy_document = policy_version['Document']
    actions_list = get_action_from_policy_document(policy_document, {'policy_id': mpolicy['PolicyId']})
    return actions_list


def get_group_members(iam_client, group_name):
    ''' get list of group members
    output: members = [{'Path': 'string','UserName': 'string', 'UserId': 'string',  'Arn': 'string',   'CreateDate': datetime(2015, 1, 1),    'PasswordLastUsed': datetime(2015, 1, 1) },]
    '''
    response = iam_client.get_group(GroupName = group_name)
    members = response['Users']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.get_group(GroupName = group_name, Marker = marker)
        members.extend(response['Users'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return members


def get_group_ipolicies(iam_client, group_name):
    ''' get list of inline policies of this group
        output: policies = ['policy1', 'policy2', ... ]
    '''
    response = iam_client.list_group_policies(GroupName = group_name)
    policies = response['PolicyNames']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_group_policies(GroupName = group_name, Marker = marker)
        policies.extend(response['PolicyNames'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return policies


def get_wuser_ipolicies(iam_client, user_name):
    ''' get list of inline policies of this user
        output: policies = ['policy1', 'policy2', ... ]
    '''
    response = iam_client.list_user_policies(UserName = user_name)
    policies = response['PolicyNames']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_user_policies(UserName = user_name, Marker = marker)
        policies.extend(response['PolicyNames'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return policies


def get_role_ipolicies(iam_client, role_name):
    ''' get list of inline policies of this role
        output: policies = ['policy1', 'policy2', ... ]
    '''
    response = iam_client.list_role_policies(RoleName = role_name)
    policies = response['PolicyNames']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_role_policies(RoleName = role_name, Marker = marker)
        policies.extend(response['PolicyNames'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return policies


def get_group_mpolicies(iam_client, group_name):
    '''
        get all managed policies that are attached to this group
        output:  policies = [ { 'PolicyName': 'string', 'PolicyArn': 'string' },]
    '''
    response = iam_client.list_attached_group_policies(GroupName = group_name)
    policies = response['AttachedPolicies']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_attached_group_policies(GroupName = group_name, Marker = marker)
        policies.extend(response['AttachedPolicies'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return policies


def get_wuser_mpolicies(iam_client, user_name):
    '''
        get all managed policies that are attached to this user
        output:  policies = [ { 'PolicyName': 'string', 'PolicyArn': 'string' },]
    '''
    response = iam_client.list_attached_user_policies(UserName = user_name)
    policies = response['AttachedPolicies']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_attached_user_policies(UserName = user_name, Marker = marker)
        policies.extend(response['AttachedPolicies'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return policies


def get_role_mpolicies(iam_client, role_name):
    '''
        get all managed policies that are attached to this role
        output:  policies = [ { 'PolicyName': 'string', 'PolicyArn': 'string' },]
    '''
    response = iam_client.list_attached_role_policies(RoleName = role_name)
    policies = response['AttachedPolicies']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_attached_role_policies(RoleName = role_name, Marker = marker)
        policies.extend(response['AttachedPolicies'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return policies


def get_wuser_access_keys(iam_client, username):
    '''
        get all managed policies that are attached to this group
        output:  user_access_keys = [{'UserName': 'string', 'AccessKeyId': 'string', 'Status': 'Active'|'Inactive','CreateDate': datetime(2015, 1, 1)},],
    '''
    response = iam_client.list_access_keys(UserName = username)
    user_access_keys = response['AccessKeyMetadata']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_access_keys(UserName = username, Marker = marker)
        user_access_keys.extend(response['AccessKeyMetadata'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return user_access_keys


def get_group_ipolicies_actions(iam_client, group_name, group_id, policy_name):
    ''' get all actions of one inline policy of a group using boto get_group_policy
        OUTPUT:  actions_list = [{ 'group_id': group_id, 'policy_name': policy_name, 'type': key, 'resource': resource, 'effect':effect, 'action': value, 'sid' : sid, 'condition': condition}]
    '''
    import json
    response = iam_client.get_group_policy(GroupName = group_name, PolicyName = policy_name)
    policy_document = response['PolicyDocument']
    actions_list = get_action_from_policy_document(policy_document, {'group_id': group_id, 'policy_name': policy_name })
    return actions_list


def get_wuser_ipolicies_actions(iam_client, user_name, user_id, policy_name):
    ''' get all actions of one inline policy of a user using boto get_user_policy
        OUTPUT:  actions_list = [{ 'wuser_id': user_id, 'policy_name': policy_name, 'type': key, 'resource': resource, 'effect':effect, 'action': value, 'sid' : sid, 'condition': condition}]
    '''
    import json
    response = iam_client.get_user_policy(UserName = user_name, PolicyName = policy_name)
    policy_document = response['PolicyDocument']
    actions_list = get_action_from_policy_document(policy_document, {'wuser_id': user_id, 'policy_name': policy_name })
    return actions_list


def get_role_ipolicies_actions(iam_client, role_name, role_id, policy_name):
    ''' get all actions of one inline policy of a role using boto get_role_policy
        OUTPUT:  actions_list = [{ 'role_id': role_id, 'policy_name': policy_name, 'type': key, 'resource': resource, 'effect':effect, 'action': value, 'sid' : sid, 'condition': condition}]
    '''
    import json
    response = iam_client.get_role_policy(RoleName = role_name, PolicyName = policy_name)
    policy_document = response['PolicyDocument']
    actions_list = get_action_from_policy_document(policy_document, {'role_id': role_id, 'policy_name': policy_name })
    return actions_list


def get_dusers_and_dusers_roles(roles_list):
    ''' extract the ad users and their roles from the Adusers excel file that is generated from AD Server
        output: dusers_list = {bensil : {'bensil': bensil, 'display_name': display_name, 'email': email, }
                dusers_roles_list = [ {'bensil': bensil, 'role_id':role_id}, ...]
    '''
    import openpyxl
    import os.path
    dusers_list = dict()
    dusers_roles_list = []
    # get the path of the current directory
    here = os.path.abspath(os.path.dirname(__file__))
    ad_users_file = AD_USERS_FILE_NAME
    # the path of the ad excel file
    path = os.path.join(here, ad_users_file)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['awsgroupexport']
    LAST_ROW = sheet.max_row
    last_cell = "D" + str(LAST_ROW)
    for cell_obj in sheet['A2':last_cell]:
        duser_inctance = {'bensil': cell_obj[0].value, 'display_name': cell_obj[1].value, 'email': cell_obj[2].value}
        dusers_list[cell_obj[0].value] = duser_inctance
        # get the account id and the name of the role from the group name(AWS-account_id-role_name, eg AWS-235651435711-ADFS-AdminAccess)
        group_name = cell_obj[3].value
        account = find_between(group_name, 'AWS-', '-')
        group_name_without_role = 'AWS-' + account + '-'
        start_index = group_name.index(group_name_without_role) + len(group_name_without_role)
        role_name = group_name[start_index:]
        # get the role id using the account id and the role name
        role_ids = [role_instance['role_id'] for role_instance in roles_list if (role_instance['role_name'] == role_name) and (role_instance['account_id'] == account)]
        role_id = role_ids[0] if len(role_ids) > 0 else None
        if role_id:
            duser_roles_inctance = {'bensil': cell_obj[0].value, 'role_id':role_id}
            dusers_roles_list.append(duser_roles_inctance)
    return dusers_list, dusers_roles_list


def get_instance_profiles(iam_client):
    '''  get all instance profiles of one account
        output: instance_profiles =  [
        {
            'Path': 'string',
            'InstanceProfileName': 'string',
            'InstanceProfileId': 'string',
            'Arn': 'string',
            'CreateDate': datetime(2015, 1, 1),
            'Roles': [
                {
                    'Path': 'string',
                    'RoleName': 'string',
                    'RoleId': 'string',
                    'Arn': 'string',
                    'CreateDate': datetime(2015, 1, 1),
                    'AssumeRolePolicyDocument': 'string',
                    'Description': 'string',
                    'MaxSessionDuration': 123
                },
            ]
    '''
    response = iam_client.list_instance_profiles()
    instance_profiles = response['InstanceProfiles']
    is_truncated = response['IsTruncated']
    if is_truncated:
        marker = response['Marker']
    while is_truncated:
        response = iam_client.list_instance_profiles(Marker = marker)
        instance_profiles.extend(response['InstanceProfiles'])
        is_truncated = response['IsTruncated']
        if is_truncated:
            marker = response['Marker']
    return instance_profiles


##############################################################################
def export_xls(workbook, worksheet_name, fieldtitles, items_list):
    ''' create excel sheet and write the list of dictionaries in it '''
    sheet = workbook.create_sheet(worksheet_name)
    xls_write(fieldtitles, items_list, sheet)


def xls_write(fieldtitles, items_list, worksheet):
    ''' write the items list in the excel sheet '''
    column_index = 1
    for title in fieldtitles :
        cell = worksheet.cell(row = 1, column = column_index)
        cell.value = title
        column_index = column_index + 1
    row_index = 2
    for item in items_list:
        column_index = 1
        for field in fieldtitles :
            cell = worksheet.cell(row = row_index, column = column_index)
            try:
                cell.value = item[field]
            except Exception:
                cell.value = ""
                print "String Error: ", field, ':', item[field]
            column_index = column_index + 1
        row_index = row_index + 1


def get_policy_id(iam_client, policy_arn):
    ''' get policy id using policy arn '''
    response = iam_client.get_policy(PolicyArn = policy_arn)
    policy_id = response['Policy']['PolicyId']
    return policy_id


def __create_db(accounts):
    ''' collect the whole information and write it in the excel file '''
    import openpyxl
    import json
    import os
    # input username and password to login aws
    username, password = get_username_password()
    # assign the current directory as the end excel report directory
    iam_reports_dir = os.path.abspath(os.path.dirname(__file__))
    accounts_list = []
    roles_list = []
    roles_ipolicies_list = []
    roles_ipolicies_actions_list = []
    roles_mpolicies_list = []

    wusers_list = []
    wusers_ipolicies_list = []
    wusers_ipolicies_actions_list = []
    wusers_mpolicies_list = []
    wusers_access_keys_list = []

    groups_list = []
    groups_ipolicies_list = []
    groups_ipolicies_actions_list = []
    groups_mpolicies_list = []
    groups_users_list = []

    mpolicies_list = {}
    mpolicies_actions_list = []

    instance_profiles_list = []
    instance_profiles_roles_list = []

    # loop over all accounts that by user inputed and collect the information and put it in lists
    for account_id, role in accounts.items():
        # login aws an get boto iam client
        iam_client = get_iam_client(username, password, account_id, role)
        # get the name of the logged-in account and add the account info to accounts list
        accounts_list.append(get_account_info(iam_client, account_id))

        # get all rolesin the logged-in account
        roles = get_roles(iam_client)
        # loop over all roles and collect their information, inline and managed policies,
        for role in roles:
            role_name = role['RoleName']
            role_id = role['RoleId']
            role_ipolicies = get_role_ipolicies(iam_client, role_name)
            for policy in role_ipolicies:
                role_ipolicy_instance = {"role_id": role_id, "policy_name": policy}
                roles_ipolicies_list.append(role_ipolicy_instance)
                role_ipolicy_actions = get_role_ipolicies_actions(iam_client, role_name, role_id, policy)
                roles_ipolicies_actions_list.extend(role_ipolicy_actions)

            role_mpolicies = get_role_mpolicies(iam_client, role_name)
            for role_mpolicy in role_mpolicies:
                mpolicy_id = get_policy_id(iam_client, role_mpolicy['PolicyArn'])
                role_mpolicy_instance = {'mpolicy_id': mpolicy_id , 'role_id': role_id }
                roles_mpolicies_list.append(role_mpolicy_instance)
            role_instance = {"role_id" : role['RoleId'], "role_name" : role['RoleName'], "role_arn" : role['Arn'] , "account_id": account_id}
            roles_list.append(role_instance)

        # get all iam users in the logged-in account
        wusers = get_wusers(iam_client)
        # loop over all iam users and collect their information, inline, managed policies, access keys
        for wuser in wusers:
            user_name = wuser['UserName']
            user_id = wuser['UserId']
            wuser_ipolicies = get_wuser_ipolicies(iam_client, user_name)
            for policy in wuser_ipolicies:
                wuser_ipolicy_instance = {"wuser_id": user_id, "policy_name": policy}
                wusers_ipolicies_list.append(wuser_ipolicy_instance)
                wuser_ipolicy_actions = get_wuser_ipolicies_actions(iam_client, user_name, user_id, policy)
                wusers_ipolicies_actions_list.extend(wuser_ipolicy_actions)

            wuser_mpolicies = get_wuser_mpolicies(iam_client, user_name)
            for wuser_mpolicy in wuser_mpolicies:
                mpolicy_id = get_policy_id(iam_client, wuser_mpolicy['PolicyArn'])
                wuser_mpolicy_instance = {'mpolicy_id': mpolicy_id , 'wuser_id': user_id }
                wusers_mpolicies_list.append(wuser_mpolicy_instance)
            wuser_access_keys = get_wuser_access_keys(iam_client, user_name)

            for access_key in wuser_access_keys:
                access_key_id = access_key['AccessKeyId']
                last_used_metadata = iam_client.get_access_key_last_used(AccessKeyId = access_key_id)
                last_used = last_used_metadata['AccessKeyLastUsed']['LastUsedDate'] if 'LastUsedDate' in last_used_metadata['AccessKeyLastUsed'].keys() else None
                access_key_instance = {"user_id" : user_id, 'access_key_id': access_key_id, 'last_used': last_used }
                wusers_access_keys_list.append(access_key_instance)

            wuser_instance = {"user_id": wuser['UserId'], "user_name": wuser['UserName'], "user_arn": wuser['Arn'], "account_id": account_id}
            wusers_list.append(wuser_instance)

        # get all groups in the logged-in account
        groups = get_groups(iam_client)

        # loop over all groups and collect their information, inline, managed policies, group members
        for group in groups:
            group_name = group['GroupName']
            group_id = group['GroupId']
            group_ipolicies = get_group_ipolicies(iam_client, group_name)
            for policy in group_ipolicies:
                group_ipolicy_instance = {"group_id": group_id, "policy_name": policy}
                groups_ipolicies_list.append(group_ipolicy_instance)
                group_ipolicy_actions = get_group_ipolicies_actions(iam_client, group_name, group_id, policy)
                groups_ipolicies_actions_list.extend(group_ipolicy_actions)

            group_mpolicies = get_group_mpolicies(iam_client, group_name)
            for group_mpolicy in group_mpolicies:
                mpolicy_id = get_policy_id(iam_client, group_mpolicy['PolicyArn'])
                group_mpolicy_instance = {'mpolicy_id': mpolicy_id , 'group_id': group_id }
                groups_mpolicies_list.append(group_mpolicy_instance)
            group_members = get_group_members(iam_client, group_name)
            for group_member in group_members:
                group_member_instance = {'user_id' : group_member['UserId'], 'group_id': group_id}
                groups_users_list.append(group_member_instance)
            group_instance = {"group_id": group['GroupId'], "group_name": group['GroupName'], "group_arn": group['Arn'], "account_id": account_id}
            groups_list.append(group_instance)

        # get all managed policies in the logged-in account
        mpolicies = get_mpolicies(iam_client)

        # loop over all managed policies and collect their information, actions
        for mpolicy in mpolicies:
            if mpolicy['PolicyId'] in mpolicies_list.keys():
                continue
            mpolicy_instance = {"policy_id": mpolicy['PolicyId'], "policy_name": mpolicy['PolicyName'], "policy_arn" : mpolicy['Arn'], "attachment_count": mpolicy['AttachmentCount']}
            mpolicies_list[mpolicy['PolicyId']] = mpolicy_instance
            mpolicy_actions_list = get_mpolicy_actions(iam_client, mpolicy)
            mpolicies_actions_list.extend(mpolicy_actions_list)

        # get all instance profiles in the logged-in account
        instance_profiles = get_instance_profiles(iam_client)

        # loop over all instance profiles and collect their information, the relationships between them and their roles
        for instance_profile in instance_profiles:
            instance_profile_instance = {'profile_id': instance_profile['InstanceProfileId'], 'profile_name': instance_profile['InstanceProfileName'], 'profile_arn': instance_profile['Arn'], 'profile_path': instance_profile['Path'], 'created_date': instance_profile['CreateDate']}
            instance_profiles_list.append(instance_profile_instance)
            instance_profile_roles = instance_profile['Roles']

            for instance_profile_role in instance_profile_roles:
                instance_profile_role_instance = {'instance_profile_id': instance_profile['InstanceProfileId'], 'role_id': instance_profile_role['RoleId']}
                instance_profiles_roles_list.append(instance_profile_role_instance)

    # collect the ad users information
    dusers_list, dusers_roles_list = get_dusers_and_dusers_roles(roles_list)
    # create excel workbook
    workbook = openpyxl.Workbook()

    # write the whole information in the excel file in different sheets
    export_xls(workbook, "accounts", ["account_id", "account_name"], accounts_list)
    export_xls(workbook, "roles", ["role_id", "role_name", "role_arn", "account_id"], roles_list)
    export_xls(workbook, "wusers", ["user_id", "user_name", "user_arn", "account_id"], wusers_list)
    export_xls(workbook, "groups", ["group_id", "group_name", "group_arn", "account_id"], groups_list)
    export_xls(workbook, "groups_ipolicies", ["group_id", "policy_name"], groups_ipolicies_list)
    export_xls(workbook, "wusers_ipolicies", ["wuser_id", "policy_name"], wusers_ipolicies_list)
    export_xls(workbook, "roles_ipolicies", ["role_id", "policy_name"], roles_ipolicies_list)
    export_xls(workbook, "groups_ipolicies_actions", ["group_id", "policy_name", 'type', 'resource', 'effect', 'action', 'sid', 'condition'], groups_ipolicies_actions_list)
    export_xls(workbook, "wusers_ipolicies_actions", ["wuser_id", "policy_name", 'type', 'resource', 'effect', 'action', 'sid', 'condition' ], wusers_ipolicies_actions_list)
    export_xls(workbook, "roles_ipolicies_actions", ["role_id", "policy_name", 'type', 'resource', 'effect', 'action', 'sid', 'condition' ], roles_ipolicies_actions_list)
    export_xls(workbook, "mpolicies", ["policy_id" , "policy_name" , "policy_arn" , "attachment_count"], mpolicies_list.values())
    export_xls(workbook, "groups_mpolicies", ['mpolicy_id' , 'group_id'], groups_mpolicies_list)
    export_xls(workbook, "wusers_mpolicies", ['mpolicy_id' , 'wuser_id'], wusers_mpolicies_list)
    export_xls(workbook, "roles_mpolicies", ['mpolicy_id' , 'role_id'], roles_mpolicies_list)
    export_xls(workbook, "mpolicies_actions", ["policy_id", 'type', 'resource', 'effect', 'action', 'sid', 'condition'], mpolicies_actions_list)
    export_xls(workbook, "dusers", ['bensil', 'display_name', 'email'], dusers_list.values())
    export_xls(workbook, "dusers_roles", ['bensil', 'role_id'], dusers_roles_list)
    export_xls(workbook, "wusers_keys", ["user_id", 'access_key_id', 'last_used'], wusers_access_keys_list)
    export_xls(workbook, "groups_users", ['user_id', 'group_id'], groups_users_list)
    export_xls(workbook, "instance_profiles", ['profile_id', 'profile_name', 'profile_arn', 'profile_path', 'created_date'], instance_profiles_list)
    export_xls(workbook, "instance_profiles_roles", ["instance_profile_id", 'role_id'], instance_profiles_roles_list)

    # save excel file
    workbook.save(os.path.join(iam_reports_dir, 'iam_db.xlsx'))


def create_db(args):
    ''' the public function of _create_db
        input: args = Namespace(accounts = {account_id: role_name, ....})
    '''

    import json
    try:
        accounts = json.loads(args.accounts)
    except Exception as er:
        print "Accounts input error: accounts should look like this: ", '{"12312312313":"ADFS-Audit"}'
        return
    __create_db(accounts)


def main():
    import json
    accounts = json.loads('{"966497653753":"ADFS-Audit", "303747409146":"ADFS-Audit", "784550693460":"ADFS-Audit", "430275495911":"ADFS-Audit", "874233888769":"ADFS-Audit", "284894803213":"ADFS-Audit", "712982040880":"ADFS-Audit" }')
    __create_db(accounts)


if __name__ == "__main__": main()
